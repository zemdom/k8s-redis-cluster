#!/usr/bin/python3

# SPDX-License-Identifier: BSD-3-Clause
# Copyright 2021-2022, Intel Corporation

import os
import sys
from argparse import ArgumentParser
from os.path import basename, dirname, isfile, join

import openpyxl


### constants ###
EMON_LABELS_COLUMN_INDEX = 'A'
EMON_AVG_VALUES_COLUMN_INDEX = 'B'
SINGLE_CLUSTER_COLUMNS = [
    'master1',
    'master2',
    'master3',
    'replica1',
    'replica2',
    'replica3'
]
#################


def parse_arguments():
    parser = ArgumentParser(
        description='Excel generator aggregating data collected by Emon monitor while benchmarking redis-cluster(s)'
    )
    parser.add_argument(
        '--pmem-generation', dest='pmem_generation', default='2', required=False,
        help='Generation of PMem dimms present in the system - possible values: 1 or 2'
    )
    parser.add_argument(
        '-p', '--paths', dest='paths', nargs='+', default=[], 
        help='Paths to directories with *.xlsx files to compare'
    )
    parser.add_argument(
        '-n', '--name', dest='result_file_name', default='results', 
        help='Name of the aggregated results file'
    )
    parser.add_argument(
        '--multiple-clusters-hostnames', dest='multiple_clusters_hostnames', nargs='+', default=[], required=False,
        help='Hostnames of k8s compute nodes - needed when creating comparison of servers instead of redis-cluster instances'
    )
    args = vars(parser.parse_args())

    return args


def extract_arguments(args):
    pmem_generation = int(args['pmem_generation'])
    paths = args['paths']
    columns = args['multiple_clusters_hostnames'] if args['multiple_clusters_hostnames'] else SINGLE_CLUSTER_COLUMNS
    result_file_name = args['result_file_name'] + '.xlsx'

    return pmem_generation, paths, columns, result_file_name


def get_emon_labels(template_name):
    template_workbook = openpyxl.load_workbook(template_name)
    template_sheet = template_workbook.active

    emon_labels = [cell.value for cell in template_sheet[EMON_LABELS_COLUMN_INDEX]]
    return emon_labels


def parse_emon_workbook(workbook_name, emon_labels):
    emon_params = [workbook_name]

    current_workbook = openpyxl.load_workbook(workbook_name)
    current_sheet = current_workbook.active

    current_sheet_emon_labels_column = current_sheet['A']

    for emon_label in emon_labels:
        for cell in current_sheet_emon_labels_column:
            if cell.value == emon_label:
                cell_number = EMON_AVG_VALUES_COLUMN_INDEX + str(cell.row)
                cell_value = current_sheet[cell_number].value
                emon_params.append(cell_value)

    return emon_params


def fill_in_workbook(source_workbook_file, emon_labels, columns, destination_workbook_sheet, run_index, all_runs_count):
    emon_params = parse_emon_workbook(source_workbook_file, emon_labels)

    for column_index, column_type in enumerate(columns):
        if column_type in source_workbook_file.lower():
            dest_column_index = (column_index * all_runs_count) + column_index + run_index + 2

            for row_index, cell_value in enumerate(emon_params):
                destination_workbook_sheet.cell(row=row_index+1, column=dest_column_index).value = cell_value


def main():
    args = parse_arguments()
    pmem_generation, paths, columns, result_file_name = extract_arguments(args)

    cpu_codename = 'CLX' if pmem_generation == 1 else 'ICX'
    template_name = f'EMON_TEMPLATE_redis-cluster_{cpu_codename}.xlsx'
    template_name = join(dirname(sys.argv[0]), 'emon_templates', template_name)

    emon_labels = get_emon_labels(template_name)
    results_workbook = openpyxl.load_workbook(template_name)
    results_sheet = results_workbook.active

    directories = [[join(path, obj) for obj in os.listdir(path)] for path in paths]
    directories = [[obj for obj in directory if isfile(obj) and obj.endswith('.xlsx')] for directory in directories]
    all_runs_count = len(directories) - 1

    for run_index, directory_content in enumerate(directories):
        for workbook_file in directory_content:
            print(f'Processing: {basename(workbook_file)}')
            fill_in_workbook(workbook_file, emon_labels, columns, results_sheet, run_index, all_runs_count)

    results_workbook.save(filename=result_file_name)
    print(f'Saved output file: {result_file_name}')


if __name__ == '__main__':
    main()
